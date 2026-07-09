from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.errors import ProcessingError
from app.models import CarrierMappingRow, Severity, ValidationIssue

REQUIRED_HEADERS = ("货代单号", "WFS Shipment ID")


def _fail(
    rule: str,
    message: str,
    *,
    filename: str | None = None,
    actual: object | None = None,
) -> None:
    raise ProcessingError(
        (
            ValidationIssue(
                severity=Severity.STRONG,
                rule=rule,
                message=message,
                repair="请修正 Excel 文件后重新上传。",
                filename=filename,
                actual=actual,
            ),
        )
    )


def read_excel_mapping(path: Path) -> tuple[CarrierMappingRow, ...]:
    if not path.is_file():
        _fail("mapping_missing", "Excel mapping file is missing", filename=path.name)

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - openpyxl error detail is library specific
        _fail(
            "mapping_unreadable",
            "Excel mapping file cannot be read",
            filename=path.name,
            actual=str(exc),
        )

    if not workbook.sheetnames:
        _fail("mapping_empty", "Excel workbook has no worksheets", filename=path.name)

    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)

    try:
        header_row = next(rows)
    except StopIteration:
        _fail("missing_headers", "Excel mapping file is empty", filename=path.name)

    headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
    if tuple(headers[:2]) != REQUIRED_HEADERS:
        _fail(
            "missing_required_headers",
            "Excel mapping file is missing required headers",
            filename=path.name,
            actual=headers[:2],
        )

    mappings: list[CarrierMappingRow] = []
    for row_number, row in enumerate(rows, start=2):
        carrier_cell = row[0] if len(row) > 0 else None
        shipment_cell = row[1] if len(row) > 1 else None
        carrier_number = str(carrier_cell).strip() if carrier_cell is not None else ""
        shipment_id = str(shipment_cell).strip() if shipment_cell is not None else ""
        if not carrier_number and not shipment_id:
            continue
        if not carrier_number or not shipment_id:
            _fail(
                "incomplete_mapping_row",
                "Excel mapping row is incomplete",
                filename=path.name,
                actual={
                    "row": row_number,
                    "carrier_number": carrier_number,
                    "shipment_id": shipment_id,
                },
            )
        mappings.append(
            CarrierMappingRow(
                row_number=row_number,
                carrier_number=carrier_number,
                shipment_id=shipment_id,
            )
        )

    return tuple(mappings)
